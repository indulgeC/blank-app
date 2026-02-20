#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build a "Crash List" table from one or many Florida Traffic Crash Report PDFs,
then append matching fields from an event.csv (by Report Number) for side-by-side comparison.

Inputs
------
1) One PDF, many PDFs, or a directory of PDFs (recursively).
2) event.csv (optional but recommended)

Output
------
- Excel (.xlsx) formatted similarly to the provided Crash List
- Optional CSV (.csv)

Notes
-----
- For coded fields (Crash Type, Severity, Light/Weather/Surface), this script outputs labels only (no leading numeric codes).
- Crash Number is generated sequentially within each Crash Year after sorting by crash date/time.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# Mappings / normalization
# ----------------------------

# User-provided Manner Of Collision codes (simplified to match Crash List style)
COLLISION_MAP = {
    1: "Rear End",
    2: "Front to Front",
    3: "Angle",
    4: "Sideswipe",
    5: "Sideswipe",
    6: "Rear to Side",
    7: "Rear to Rear",
    77: "Other",
    88: "Unknown",
}

# User-provided Injury Severity codes, simplified to match Crash List style
# 1 -> No Injury
# 2-3 -> Injury
# 4 -> Serious Injury
# 5-6 -> Fatal

# 1 No Injury
# 2 Possible Injury
# 3 Non‐incapacitating Injury
# 4 Incapacitating Injury
# 5 Fatal (within 30 days)
# 6 Non‐Traffic Fatality

def injury_simplify(code: int) -> str:
    # if code == 1:
    #     return "No Injury"
    # if code in (2, 3):
    #     return "Injury"
    # if code == 4:
    #     return "Serious Injury"
    # if code in (5, 6):
    #     return "Fatal"
    # return ""

    if code == 1:
        return "No Injury"
    if code == 2:
        return "Possible Injury"
    if code == 3:
        return "Non‐incapacitating Injury"
    if code == 4:
        return "Incapacitating Injury"
    if code == 5:
        return "Fatal (within 30 days)"
    if code == 6:
        return "Non‐Traffic Fatality"
    return ""

def normalize_condition(s: str) -> str:
    """Standardize common label variants, and keep only the label (no numeric code)."""
    if not s:
        return ""
    s = str(s).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s)

    # Normalize common light variants
    s = re.sub(r"(?i)\bdark-?lighted\b", "Dark - Lighted", s)
    s = re.sub(r"(?i)\bdark\s*-\s*lighted\b", "Dark - Lighted", s)
    s = re.sub(r"(?i)\bdark-?\s*not-?\s*lighted\b", "Dark - Not Lighted", s)
    s = re.sub(r"(?i)\bdark\s*-\s*not\s*lighted\b", "Dark - Not Lighted", s)

    return s


def strip_leading_code(value: str) -> str:
    """
    If a value looks like '4 Dark-Lighted' or '3 Rain', strip the leading numeric code.
    Safe for coded fields, but do NOT use this for Date strings.
    """
    if value is None:
        return ""
    value = str(value).strip()
    value = re.sub(r"^\s*\d{1,3}\s*[-:]?\s*", "", value)
    return value.strip()


# ----------------------------
# PDF parsing
# ----------------------------

def pdf_to_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse_date_of_crash(text: str) -> Optional[dt.datetime]:
    m = re.search(
        r"Date of Crash\s*\n\s*([0-9]{1,2}/[A-Za-z]{3}/[0-9]{4})\s+([0-9]{1,2}:[0-9]{2})\s*([AP]M)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        return None
    s = f"{m.group(1)} {m.group(2)} {m.group(3)}"
    try:
        return dt.datetime.strptime(s, "%d/%b/%Y %I:%M %p")
    except Exception:
        return None


def parse_report_number(text: str) -> str:
    m = re.search(r"HSMV Crash Report Number\s*\n\s*([0-9]{6,10})", text, flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""


def parse_code_and_label(text: str, section_label: str) -> (Optional[int], str):
    """
    Parse blocks like:
      Weather Condition
      3 Rain
    """
    m = re.search(
        rf"{re.escape(section_label)}\s*\n\s*([0-9]{{1,3}})\s+([^\n\r]+)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        return None, ""
    code = int(m.group(1))
    label = m.group(2).strip()
    return code, label


def parse_injury_severity_codes(text: str) -> List[int]:
    """
    Find all occurrences of:
      Injury Severity
      <code> <label>
    Then return all codes (we will take the worst / max).
    """
    matches = re.findall(
        r"Injury Severity\s*\n\s*([0-9]{1,2})\s+[^\n\r]+",
        text,
        flags=re.IGNORECASE,
    )
    out: List[int] = []
    for c in matches:
        try:
            out.append(int(c))
        except Exception:
            pass
    return out


def extract_one_pdf(pdf_path: Path) -> Dict[str, object]:
    text = pdf_to_text(pdf_path)

    crash_dt = parse_date_of_crash(text)
    crash_year = crash_dt.year if crash_dt else ""

    report_num = parse_report_number(text)

    light_code, light_label = parse_code_and_label(text, "light Condition")
    weather_code, weather_label = parse_code_and_label(text, "Weather Condition")
    surface_code, surface_label = parse_code_and_label(text, "Roadway Surface Condition")
    moc_code, moc_label = parse_code_and_label(text, "Manner Of Collision")

    injury_codes = parse_injury_severity_codes(text)
    worst_injury = max(injury_codes) if injury_codes else None

    return {
        "Crash Year": crash_year,
        "Crash Number": None,  # fill later
        "Report Number": report_num,
        "Date": crash_dt,  # keep as datetime for sorting
        "Crash Type": COLLISION_MAP.get(moc_code, strip_leading_code(moc_label) or ""),
        "Severity": injury_simplify(worst_injury) if worst_injury is not None else "",
        "Light": normalize_condition(strip_leading_code(light_label)),
        "Weather": normalize_condition(strip_leading_code(weather_label)),
        "Surface Condition": normalize_condition(strip_leading_code(surface_label)),
        "_src_pdf": str(pdf_path),
    }


# ----------------------------
# event.csv parsing (by Excel-like column letters)
# ----------------------------

def excel_col_to_index(col: str) -> int:
    col = col.upper()
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def parse_event_datetime(s: str) -> Optional[dt.datetime]:
    if s is None:
        return None
    s = str(s).strip()
    fmts = [
        "%d-%b-%Y %I:%M %p",   # 18-JUN-2024 02:52 PM
        "%m/%d/%Y %H:%M",      # 2/15/2020 23:14
        "%m/%d/%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
    ]
    for f in fmts:
        try:
            return dt.datetime.strptime(s, f)
        except Exception:
            pass
    try:
        v = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(v) else v.to_pydatetime()
    except Exception:
        return None


def load_event_lookup(event_csv: Path) -> Dict[str, Dict[str, str]]:
    """
    User mapping:
      Report Number -> column A
      Crash Year    -> column B
      Date          -> column C
      Light         -> column AF
      Weather       -> column AG
      Surface       -> column AH
      Crash Type    -> column BH
      Severity      -> column BI
    """
    df = pd.read_csv(event_csv, dtype=str, low_memory=False)

    cols = list(df.columns)
    def col(letter: str) -> str:
        return cols[excel_col_to_index(letter)]

    out = df[
        [col("B"), col("A"), col("C"), col("BH"), col("BI"), col("AF"), col("AG"), col("AH")]
    ].copy()
    out.columns = [
        "Crash Year",
        "Report Number",
        "Date",
        "Crash Type",
        "Severity",
        "Light",
        "Weather",
        "Surface Condition",
    ]

    out["Report Number"] = out["Report Number"].astype(str).str.strip()

    # Normalize date to "m/d/YYYY HH:MM" when possible
    out["_dt"] = out["Date"].apply(parse_event_datetime)
    out["Date"] = out["_dt"].apply(
        lambda d: d.strftime("%-m/%-d/%Y %H:%M") if isinstance(d, dt.datetime) else ""
    )
    out.drop(columns=["_dt"], inplace=True)

    # Clean coded fields (just in case)
    for c in ["Crash Type", "Severity", "Light", "Weather", "Surface Condition"]:
        out[c] = out[c].fillna("").map(lambda x: normalize_condition(strip_leading_code(str(x))))

    lookup: Dict[str, Dict[str, str]] = {}
    for _, r in out.iterrows():
        rn = str(r["Report Number"]).strip()
        if rn and rn not in lookup:
            lookup[rn] = r.to_dict()

    return lookup


# ----------------------------
# Output formatting
# ----------------------------

def format_dt_out(d: object) -> str:
    if isinstance(d, dt.datetime):
        return d.strftime("%-m/%-d/%Y %H:%M")
    return ""


def build_dataframe(pdf_rows: List[Dict[str, object]], event_lookup: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    # Sort by (Crash Year, Date)
    pdf_rows = sorted(
        pdf_rows,
        key=lambda r: (r.get("Crash Year") or 0, r.get("Date") or dt.datetime.min),
    )

    # Assign Crash Number sequentially within each year
    counters: Dict[int, int] = {}
    for r in pdf_rows:
        y = int(r["Crash Year"]) if r.get("Crash Year") else 0
        counters.setdefault(y, 0)
        counters[y] += 1
        r["Crash Number"] = counters[y]

    # Build final rows with appended event fields
    out_rows = []
    for r in pdf_rows:
        base = {
            "Crash Year": r.get("Crash Year", ""),
            "Crash Number": r.get("Crash Number", ""),
            "Report Number": r.get("Report Number", ""),
            "Date": format_dt_out(r.get("Date")),
            "Crash Type": r.get("Crash Type", ""),
            "Severity": r.get("Severity", ""),
            "Light": r.get("Light", ""),
            "Weather": r.get("Weather", ""),
            "Surface Condition": r.get("Surface Condition", ""),
        }

        ev = event_lookup.get(str(base["Report Number"]).strip())
        if ev:
            base.update({
                "Crash Year (event)": ev.get("Crash Year", ""),
                "Report Number (event)": ev.get("Report Number", ""),
                "Date (event)": ev.get("Date", ""),
                "Crash Type (event)": ev.get("Crash Type", ""),
                "Severity (event)": ev.get("Severity", ""),
                "Light (event)": ev.get("Light", ""),
                "Weather (event)": ev.get("Weather", ""),
                "Surface Condition (event)": ev.get("Surface Condition", ""),
            })
        else:
            base.update({
                "Crash Year (event)": "",
                "Report Number (event)": "",
                "Date (event)": "",
                "Crash Type (event)": "",
                "Severity (event)": "",
                "Light (event)": "",
                "Weather (event)": "",
                "Surface Condition (event)": "",
            })

        out_rows.append(base)

    return pd.DataFrame(out_rows)


# Column groups for group header row ("Police Report" / "Event")
_POLICE_COLS = ["Crash Year", "Crash Number", "Report Number", "Date", "Crash Type", "Severity", "Light", "Weather", "Surface Condition"]
_EVENT_COLS = ["Crash Year (event)", "Report Number (event)", "Date (event)", "Crash Type (event)", "Severity (event)", "Light (event)", "Weather (event)", "Surface Condition (event)"]


def write_xlsx(df: pd.DataFrame, out_path: Path, sheet_name: str = "Crash List") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style to resemble provided Crash List
    header_fill = PatternFill("solid", fgColor="C00000")  # dark red
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)

    # Find column ranges for group headers
    police_start = None
    police_end = None
    event_start = None
    event_end = None
    for i, c in enumerate(cols):
        if c in _POLICE_COLS:
            if police_start is None:
                police_start = i + 1
            police_end = i + 1
        elif c in _EVENT_COLS:
            if event_start is None:
                event_start = i + 1
            event_end = i + 1

    # Row 1: group headers ("Police Report" / "Event") - white bg, black text
    group_fill = PatternFill("solid", fgColor="FFFFFF")
    group_font = Font(color="000000", bold=True)
    group_row = 1
    if police_start is not None and police_end is not None:
        cell = ws.cell(row=group_row, column=police_start, value="Police Report")
        cell.fill = group_fill
        cell.font = group_font
        cell.alignment = center
        cell.border = border
        if police_end > police_start:
            ws.merge_cells(start_row=group_row, start_column=police_start, end_row=group_row, end_column=police_end)
    if event_start is not None and event_end is not None:
        cell = ws.cell(row=group_row, column=event_start, value="Event")
        cell.fill = group_fill
        cell.font = group_font
        cell.alignment = center
        cell.border = border
        if event_end > event_start:
            ws.merge_cells(start_row=group_row, start_column=event_start, end_row=group_row, end_column=event_end)

    # Row 2: column headers
    header_row = 2
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Rows 3+: data
    data_start_row = 3
    for _, r in df.iterrows():
        ws.append([r[c] for c in cols])
    # Fix: append adds to next row, so we need to write data at correct rows
    # Actually append always adds to the next empty row. After we wrote group_row and header_row, the next append goes to row 3. Good.
    # But wait - we didn't use append for header. We used cell() for group_row and header_row. So the "next" row for append would be row 1 (first append) -> no.
    # Let me trace: we created cells at row 1 (group) and row 2 (headers). We never called append. So the sheet has row 1 and 2 filled. When we do append(), it appends to the first empty row. In openpyxl, append adds a new row at max_row + 1. Before any append, max_row might be 2. So first append adds row 3. Good.

    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=1, max_col=len(cols)):
        for cell in row:
            cell.alignment = center
            cell.border = border

    ws.freeze_panes = "A3"
    # AutoFilter on row 2 (column headers) so sort/filter dropdowns appear there
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{ws.max_row}"

    # Column widths (cap for readability)
    for j, col in enumerate(cols, start=1):
        maxlen = max(
            len(str(col)),
            len("Police Report") if police_start is not None and police_start <= j <= (police_end or 0) else 0,
            len("Event") if event_start is not None and event_start <= j <= (event_end or 0) else 0,
            *(len(str(ws.cell(row=i, column=j).value or "")) for i in range(data_start_row, min(ws.max_row, 200) + 1)),
        )
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, maxlen + 2), 28)

    wb.save(str(out_path))


# ----------------------------
# CLI
# ----------------------------

def collect_pdfs(inputs: List[str]) -> List[Path]:
    pdfs: List[Path] = []
    for p in inputs:
        path = Path(p)
        if path.is_dir():
            pdfs.extend(sorted(path.rglob("*.pdf")))
        else:
            # file or glob
            if "*" in p or "?" in p or "[" in p:
                pdfs.extend(sorted(Path().glob(p)))
            else:
                pdfs.append(path)
    # unique + keep order
    seen = set()
    out = []
    for x in pdfs:
        x = x.resolve()
        if x.exists() and x.suffix.lower() == ".pdf" and x not in seen:
            out.append(x)
            seen.add(x)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--reports",
        nargs="+",
        required=True,
        help="PDF file(s), a directory, or a glob pattern (e.g., ./reports/*.pdf).",
    )
    ap.add_argument("--event_csv", default=None, help="Path to event.csv")
    ap.add_argument("--out_xlsx", default="Crash_List_final.xlsx", help="Output Excel path")
    ap.add_argument("--out_csv", default=None, help="Optional output CSV path")
    args = ap.parse_args()

    pdf_files = collect_pdfs(args.reports)
    if not pdf_files:
        raise SystemExit("No PDF reports found. Check --reports path(s).")

    event_lookup: Dict[str, Dict[str, str]] = {}
    if args.event_csv:
        event_lookup = load_event_lookup(Path(args.event_csv))

    rows: List[Dict[str, object]] = []
    for pdf in pdf_files:
        try:
            rows.append(extract_one_pdf(pdf))
        except Exception as e:
            print(f"[WARN] Failed to parse {pdf.name}: {e}")

    df_out = build_dataframe(rows, event_lookup)
    out_xlsx = Path(args.out_xlsx)
    write_xlsx(df_out, out_xlsx)

    if args.out_csv:
        Path(args.out_csv).parent.mkdir(parents=True, exist_ok=True)
        df_out.to_csv(args.out_csv, index=False)

    print(f"Done. Wrote: {out_xlsx.resolve()}")
    if args.out_csv:
        print(f"Also wrote: {Path(args.out_csv).resolve()}")


if __name__ == "__main__":
    main()
